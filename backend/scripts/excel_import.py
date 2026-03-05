#!/usr/bin/env python3
"""
Halb-automatisierter Excel-Import für Berechnungsvorschriften.

Liest eine Excel-Datei gemäß YAML-Config, extrahiert Formelzellen aus den
konfigurierten Tabellenbereichen und erzeugt Zelleneingaben. Diese können
als Dry-Run (JSON/CSV) ausgegeben oder per API importiert werden.

Verwendung:
  # Dry-Run (Vorschau, keine API-Aufrufe):
  python scripts/excel_import.py --config config/excel_import_config.yaml --excel datei.xlsx

  # Import (API-Aufrufe):
  python scripts/excel_import.py --config config/excel_import_config.yaml --excel datei.xlsx --import

  # API-URL (Standard: http://localhost:8000):
  API_URL=http://localhost:8000 python scripts/excel_import.py ...
"""
import argparse
import json
import logging
import os
import sys
from pathlib import Path

# Backend-Verzeichnis für Imports
sys.path.insert(0, str(Path(__file__).parent.parent))

import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries, column_index_from_string
from openpyxl.worksheet.formula import ArrayFormula

from utils.formel_utils import formel_excel_normalisieren, tabellenspalten_aus_formel
from utils.excel_referenz_index import lade_referenz_index

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)


def _config_zu_triplets(config: dict) -> list:
    """
    Konfiguration in Triplets (tabellenidentifikator, blatt_name, tabelle) auflösen.

    Unterstützt zwei Strukturen:
    - Neu: tabellenidentifikatoren -> tabellenblaetter -> tabellen (id auf oberster Ebene)
    - Alt: tabellenblaetter -> tabellen (id pro Tabelle)

    Returns:
        Liste von (tabellenidentifikator, blatt_name, tabelle_config)
    """
    result = []

    # Neue Hierarchie: Tabellenidentifikator -> Tabellenblatt -> Tabellen
    for tid_block in config.get("tabellenidentifikatoren", []):
        t_id = tid_block.get("id", "Tabelle1")
        for blatt in tid_block.get("tabellenblaetter", []):
            blatt_name = blatt.get("name")
            if not blatt_name:
                logger.warning("Tabellenblatt ohne Namen übersprungen")
                continue
            for tabelle in blatt.get("tabellen", []):
                result.append((t_id, blatt_name, tabelle))

    # Fallback: alte Struktur tabellenblaetter -> tabellen (id pro Tabelle)
    if not result:
        for blatt in config.get("tabellenblaetter", []):
            blatt_name = blatt.get("name")
            if not blatt_name:
                logger.warning("Tabellenblatt ohne Namen übersprungen")
                continue
            for tabelle in blatt.get("tabellen", []):
                t_id = tabelle.get("id", "Tabelle1")
                result.append((t_id, blatt_name, tabelle))

    return result


def parse_range(bereich: str) -> tuple:
    """
    Parst Bereich wie 'A5:F16' in (min_row, min_col, max_row, max_col).
    openpyxl verwendet 1-basierte Indizes.
    """
    try:
        min_col, min_row, max_col, max_row = range_boundaries(bereich)
        return min_row, min_col, max_row, max_col
    except Exception:
        raise ValueError(f"Ungültiger Bereich: {bereich}")


def zelle_zu_ref(row: int, col: int) -> str:
    """Wandelt (row, col) in Excel-Referenz um (z.B. D7)."""
    return f"{get_column_letter(col)}{row}"


def zellenwert_mit_merge(ws, row: int, col: int):
    """
    Liest den Zellwert. Bei zusammengeführten Zellen (Merge) liefert nur die
    obere linke Zelle den Wert; MergedCell-Platzhalter sind leer. Diese Funktion
    ermittelt bei leerem Wert, ob die Zelle in einem Merge-Bereich liegt, und
    gibt in dem Fall den Wert der oberen linken Zelle zurück.
    """
    cell = ws.cell(row=row, column=col)
    val = cell.value
    if val is not None and str(val).strip():
        return val
    # Prüfen, ob Zelle Teil eines Merge-Bereichs ist – dann Wert der Top-Left
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            return top_left.value
    return val


def ist_beschreibungszelle(
    row: int, col: int, min_row: int, min_col: int, tabelle_config: dict
) -> bool:
    """
    Prüft, ob die Zelle (row, col) zu den als Beschreibung konfigurierten
    Spalten/Zeilen gehört. Solche Zellen werden nicht als Berechnungsvorschrift
    importiert (Überschriften, Spalten-/Zeilenbeschriftungen).

    Nur relevant bei beschreibung_quelle "zellen" mit beschreibung_aus_zellen:
    - erste_spalte_gleiche_zeile: erste Spalte = Beschriftung der Zeile
    - gleiche_spalte_erste_n_zeilen: erste n Zeilen = Beschriftung der Spalte
    """
    quelle = tabelle_config.get("beschreibung_quelle")
    if quelle != "zellen":
        return False

    aus_zellen = tabelle_config.get("beschreibung_aus_zellen", {})
    if not aus_zellen:
        return False

    # erste_spalte: Zellen in der ersten Spalte sind Zeilenbeschriftung
    if aus_zellen.get("erste_spalte_gleiche_zeile") and col == min_col:
        return True

    # erste n Zeilen: Zellen in den ersten n Zeilen sind Spaltenbeschriftung
    n_zeilen = aus_zellen.get("gleiche_spalte_erste_n_zeilen", 0)
    if n_zeilen and n_zeilen > 0 and row < min_row + n_zeilen:
        return True

    return False


def beschreibung_aus_zellen_ermitteln(
    ws,
    formel_row: int,
    formel_col: int,
    min_row: int,
    min_col: int,
    config: dict,
) -> str:
    """
    Ermittelt die Beschreibung aus konfigurierten Zellen relativ zum Tabellenbereich.
    config: beschreibung_aus_zellen mit erste_spalte_gleiche_zeile, gleiche_spalte_erste_n_zeilen, trennzeichen
    """
    teile = []
    trennzeichen = config.get("trennzeichen", " – ")

    # gleiche_spalte_erste_n_zeilen: Werte aus gleicher Spalte, erste n Zeilen
    n_zeilen = config.get("gleiche_spalte_erste_n_zeilen", 0)
    if n_zeilen and n_zeilen > 0:
        for r in range(min_row, min(min_row + n_zeilen, formel_row)):
            val = zellenwert_mit_merge(ws, r, formel_col)
            if val is not None and str(val).strip():
                teile.append(str(val).strip())

    # erste_spalte_gleiche_zeile: Wert aus erster Spalte, gleiche Zeile (Merge berücksichtigt)
    if config.get("erste_spalte_gleiche_zeile") and formel_col > min_col:
        val = zellenwert_mit_merge(ws, formel_row, min_col)
        if val is not None and str(val).strip():
            teile.append(str(val).strip())

    return trennzeichen.join(teile) if teile else ""


def beschreibung_ermitteln(
    ws, cell, formel_row: int, formel_col: int, min_row: int, min_col: int, tabelle_config: dict
) -> str:
    """Ermittelt die Beschreibung gemäß beschreibung_quelle der Tabelle."""
    quelle = tabelle_config.get("beschreibung_quelle", "formel")

    if quelle == "zellen":
        aus_zellen = tabelle_config.get("beschreibung_aus_zellen", {})
        if aus_zellen:
            return beschreibung_aus_zellen_ermitteln(
                ws, formel_row, formel_col, min_row, min_col, aus_zellen
            )

    if quelle == "kommentar":
        if hasattr(cell, "comment") and cell.comment:
            return (cell.comment.text or "").strip().replace("\n", " ")

    if quelle == "links" and formel_col > 1:
        val = zellenwert_mit_merge(ws, formel_row, formel_col - 1)
        if val is not None:
            return str(val).strip()

    if quelle == "oben" and formel_row > 1:
        val = zellenwert_mit_merge(ws, formel_row - 1, formel_col)
        if val is not None:
            return str(val).strip()

    # Fallback: formel oder leere Beschreibung (formel als Hinweis)
    if quelle == "formel":
        formel = cell.value or ""
        return str(formel).strip()[:200]  # Begrenzen für Lesbarkeit

    return ""


def formel_ersetzung_anwenden(beschreibung: str, formel_ersetzung: dict | None) -> str:
    """
    Ersetzt in der Beschreibung vorkommende Formeln durch die konfigurierten Texte.
    formel_ersetzung: Mapping Formel -> Ersatztext (z.B. "=$'INTERN BEZÜGE'.$D$3" -> "Vollzeit festangestellt").
    Längere Formeln werden zuerst ersetzt, um Überlappungen zu vermeiden.
    """
    if not beschreibung or not formel_ersetzung:
        return beschreibung
    # Nach Länge absteigend sortieren, damit längere Formeln vor kürzeren ersetzt werden
    for formel in sorted(formel_ersetzung.keys(), key=len, reverse=True):
        beschreibung = beschreibung.replace(formel, formel_ersetzung[formel])
    return beschreibung


def zelleneingaben_aus_excel(config_path: str, excel_path: str) -> list:
    """
    Liest die Config und Excel-Datei und erzeugt eine Liste von Zelleneingabe-Dicts.
    """
    with open(config_path, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)

    excel_datei_config = config.get("excel_datei")
    # Übersteuern, falls expliziter Pfad angegeben
    pfad = excel_path or excel_datei_config
    if not pfad or not Path(pfad).exists():
        raise FileNotFoundError(f"Excel-Datei nicht gefunden: {pfad}")

    wb = load_workbook(pfad, read_only=False, data_only=False)
    zelleneingaben = []
    # Optional: Formeln in Beschreibungen durch lesbare Texte ersetzen
    formel_ersetzung = config.get("formel_ersetzung") or {}

    # Referenz-Index vorab laden: Tabellen und benannte Bereiche aus Excel auflösen.
    # Ermöglicht bei Formeln mit MAJahr1[Spalte] direkt Blatt+Bereich einzutragen.
    # Fallback: Wenn Auflösung fehlschlägt, bleibt referenz_bereiche leer.
    referenz_index = lade_referenz_index(wb)

    # Konfiguration in Triplets (tabellenidentifikator, blatt_name, tabelle) auflösen
    # Neue Hierarchie: tabellenidentifikatoren -> tabellenblaetter -> tabellen
    # Fallback: alte Struktur tabellenblaetter -> tabellen (id pro Tabelle)
    config_triplets = _config_zu_triplets(config)

    for t_id, blatt_name, tabelle in config_triplets:
        ws = wb[blatt_name] if blatt_name in wb.sheetnames else None
        if not ws:
            logger.warning(f"Tabellenblatt '{blatt_name}' nicht gefunden, übersprungen")
            continue

        bereich = tabelle.get("bereich")
        wichtige_zellen = set(tabelle.get("wichtige_zellen") or [])

        if not bereich:
            logger.warning(f"Tabelle {t_id} ohne Bereich übersprungen")
            continue

        try:
            min_row, min_col, max_row, max_col = parse_range(bereich)
        except ValueError as e:
            logger.warning(f"Tabelle {t_id}: {e}")
            continue

        # Optional: Nur bestimmte Spalten als Formelzellen verarbeiten (z.B. formel_spalten: ["G"])
        # Ohne Angabe: alle Spalten im Bereich außer Beschreibungszellen
        formel_spalten = tabelle.get("formel_spalten")  # Liste von Buchstaben: ["G"] oder ["G", "H"]
        formel_col_indices = None
        if formel_spalten:
            formel_col_indices = set()
            for sp in formel_spalten:
                try:
                    formel_col_indices.add(column_index_from_string(str(sp).strip().upper()))
                except (ValueError, TypeError):
                    pass
            if not formel_col_indices:
                formel_col_indices = None  # Keine gültigen Spalten – Filter ignorieren

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                # Beschreibungszellen (Überschriften, Zeilen-/Spaltenbeschriftung) überspringen
                if ist_beschreibungszelle(row, col, min_row, min_col, tabelle):
                    continue
                # Optional: Nur konfigurierte Formel-Spalten verarbeiten (Rest nur für Beschreibung)
                if formel_col_indices is not None and col not in formel_col_indices:
                    continue
                cell = ws.cell(row=row, column=col)
                val = cell.value
                # Formel extrahieren: normale Zellen haben String "=...", ArrayFormulas haben .text
                if isinstance(val, ArrayFormula) and getattr(val, "text", None):
                    formel = val.text.strip() if isinstance(val.text, str) else ""
                elif val and isinstance(val, str) and val.strip().startswith("="):
                    formel = val.strip()
                else:
                    continue
                if not formel or not formel.startswith("="):
                    continue
                # Excel-interne Präfixe entfernen (_xlfn.IFS -> IFS), damit Formel lesbar bleibt
                formel = formel_excel_normalisieren(formel)
                zellen_ref = zelle_zu_ref(row, col)
                beschreibung = beschreibung_ermitteln(
                    ws, cell, row, col, min_row, min_col, tabelle
                )
                beschreibung = formel_ersetzung_anwenden(beschreibung, formel_ersetzung)
                wichtig = zellen_ref in wichtige_zellen

                ze = {
                    "tabellenidentifikator": t_id,
                    "tabellenblatt": blatt_name,
                    "zellenidentifikator": zellen_ref,
                    "beschreibung": beschreibung or formel[:50],
                    "formel": formel,
                    "wichtig": wichtig,
                }
                # Referenz-Index: Tabellenspalten aus Formel auflösen (MAJahr1[Spalte] -> Blatt+Bereich).
                # Fallback: Keine Anreicherung wenn Lookup fehlschlägt.
                tabellenspalten = tabellenspalten_aus_formel(formel)
                if tabellenspalten and referenz_index:
                    referenz_bereiche = []
                    for ts in tabellenspalten:
                        key = (ts["tabelle"], ts["spalte"])
                        resolved = referenz_index.get(key)
                        if resolved:
                            referenz_bereiche.append({
                                "tabelle": ts["tabelle"],
                                "spalte": ts["spalte"],
                                "blatt": resolved["blatt"],
                                "bereich": resolved["bereich"],
                            })
                    if referenz_bereiche:
                        ze["referenz_bereiche"] = referenz_bereiche
                zelleneingaben.append(ze)
                logger.debug(f"Zelleneingabe: {zellen_ref} ({t_id}, {blatt_name})")

    wb.close()
    return zelleneingaben


def dry_run(zelleneingaben: list, output_json: bool = True) -> None:
    """Gibt Zelleneingaben als JSON oder CSV zur Prüfung aus."""
    if output_json:
        print(json.dumps(zelleneingaben, indent=2, ensure_ascii=False))
    else:
        if not zelleneingaben:
            print("Keine Zelleneingaben")
            return
        headers = list(zelleneingaben[0].keys())
        print(",".join(headers))
        for ze in zelleneingaben:
            print(",".join(str(ze.get(h, "")).replace(",", ";") for h in headers))


def importiere_per_api(zelleneingaben: list, api_url: str) -> None:
    """
    Sendet jede Zelleneingabe per POST an die API.
    Fehlgeschlagene Zelleneingaben werden gesammelt und am Ende als JSON ausgegeben.
    """
    import urllib.request
    import urllib.error

    base = api_url.rstrip("/")
    endpoint = f"{base}/api/berechnungsvorschriften"
    erstellt = 0
    fehlgeschlagen = []  # Liste: [ {"zelleneingabe": {...}, "fehler": "..."}, ... ]

    for i, ze in enumerate(zelleneingaben):
        payload = json.dumps({"zelleneingabe": ze}).encode("utf-8")
        req = urllib.request.Request(
            endpoint,
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                if 200 <= resp.getcode() < 300:
                    erstellt += 1
                    logger.info(f"({i + 1}/{len(zelleneingaben)}) Erstellt: {ze['zellenidentifikator']} ({ze['tabellenblatt']})")
                else:
                    fehlermeldung = f"HTTP {resp.getcode()}"
                    fehlgeschlagen.append({"zelleneingabe": ze, "fehler": fehlermeldung})
                    logger.warning(f"({i + 1}/{len(zelleneingaben)}) {fehlermeldung}: {ze['zellenidentifikator']}")
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", errors="replace") if e.fp else ""
            fehlermeldung = f"HTTP {e.code}: {body[:300]}"
            fehlgeschlagen.append({"zelleneingabe": ze, "fehler": fehlermeldung})
            logger.error(f"({i + 1}/{len(zelleneingaben)}) Fehler {ze['zellenidentifikator']}: {e.code} {body[:200]}")
        except Exception as e:
            fehlermeldung = str(e)
            fehlgeschlagen.append({"zelleneingabe": ze, "fehler": fehlermeldung})
            logger.error(f"({i + 1}/{len(zelleneingaben)}) Fehler {ze['zellenidentifikator']}: {e}")

    logger.info(f"Import abgeschlossen: {erstellt} erstellt, {len(fehlgeschlagen)} Fehler")

    # Fehlgeschlagene Zelleneingaben am Ende ausgeben (für Nachbearbeitung oder erneuten Import)
    if fehlgeschlagen:
        print("\n--- Fehlgeschlagene Zelleneingaben ---", file=sys.stderr)
        print(json.dumps(fehlgeschlagen, indent=2, ensure_ascii=False), file=sys.stderr)


def main():
    parser = argparse.ArgumentParser(description="Excel-Import für Berechnungsvorschriften")
    parser.add_argument("--config", "-c", required=True, help="Pfad zur YAML-Config")
    parser.add_argument("--excel", "-e", default=None, help="Pfad zur Excel-Datei (überschreibt Config)")
    parser.add_argument("--import", dest="do_import", action="store_true", help="Import per API ausführen")
    parser.add_argument("--csv", action="store_true", help="Dry-Run als CSV statt JSON")
    parser.add_argument("--api-url", default=None, help="API-Basis-URL (default: API_URL oder localhost:8000)")
    args = parser.parse_args()

    config_path = Path(args.config)
    if not config_path.is_absolute():
        config_path = Path(__file__).parent.parent / config_path
    if not config_path.exists():
        logger.error(f"Config nicht gefunden: {config_path}")
        sys.exit(1)

    excel_path = args.excel
    if excel_path and not Path(excel_path).is_absolute():
        # Relativ zum Projekt-Root oder CWD
        excel_path = str(Path(excel_path).resolve())

    try:
        zelleneingaben = zelleneingaben_aus_excel(str(config_path), excel_path)
    except FileNotFoundError as e:
        logger.error(str(e))
        sys.exit(1)
    except Exception as e:
        logger.exception(f"Fehler beim Lesen: {e}")
        sys.exit(1)

    logger.info(f"{len(zelleneingaben)} Zelleneingaben extrahiert")

    if args.do_import:
        api_url = args.api_url or os.environ.get("API_URL", "http://localhost:8000")
        importiere_per_api(zelleneingaben, api_url)
    else:
        dry_run(zelleneingaben, output_json=not args.csv)


if __name__ == "__main__":
    main()
