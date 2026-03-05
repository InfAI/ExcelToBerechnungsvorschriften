"""
Excel-Referenz-Index: Auflösung von Tabellen- und Bereichsreferenzen aus Excel-Dateien.

Lädt alle referenzierbaren Tabellen und benannten Bereiche vorab, damit beim
Auftauchen in einer Formel (z.B. MAJahr1[Angestelltenverhältnis]) direkt der
aufgelöste Blatt+Bereich eingetragen werden kann.

Auflösungsreihenfolge:
1. ws.tables (ListObjects) – volle Spaltenstruktur
2. wb.defined_names – globale benannte Bereiche
3. ws.defined_names – blattspezifische benannte Bereiche
"""
import logging
from typing import Any, Optional

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

logger = logging.getLogger(__name__)


def _spaltenbereich_aus_tabelle(
    table_ref: str,
    table_columns: list,
    spaltenname: str,
    sheet_name: str,
) -> Optional[dict[str, str]]:
    """
    Berechnet den Zellbereich für eine Spalte innerhalb einer Tabelle.

    Args:
        table_ref: table.ref (z.B. "B5:G20")
        table_columns: table.tableColumns (Liste)
        spaltenname: Name der Spalte (z.B. "Angestelltenverhältnis")
        sheet_name: Name des Blatts

    Returns:
        {"blatt": "2. Arbeitszeit AW", "bereich": "C6:C20", "bereich_mit_blatt": "'2. Arbeitszeit AW'!C6:C20"}
        oder None wenn Spalte nicht gefunden
    """
    try:
        min_col, min_row, max_col, max_row = range_boundaries(table_ref)
    except Exception as e:
        logger.debug(f"range_boundaries fehlgeschlagen für {table_ref}: {e}")
        return None

    for i, col in enumerate(table_columns):
        col_name = getattr(col, "name", None) or ""
        if str(col_name).strip() == str(spaltenname).strip():
            spalten_col = min_col + i
            # Datenbereich: Header-Zeile (min_row) überspringen, ab min_row+1
            bereich = f"{get_column_letter(spalten_col)}{min_row + 1}:{get_column_letter(spalten_col)}{max_row}"
            bereich_mit_blatt = f"'{sheet_name}'!{bereich}"
            return {
                "blatt": sheet_name,
                "bereich": bereich,
                "bereich_mit_blatt": bereich_mit_blatt,
            }
    return None


def lade_referenz_index(wb: Any) -> dict[tuple[str, Optional[str]], dict[str, str]]:
    """
    Lädt alle Tabellen- und Bereichsreferenzen aus dem Workbook.

    Returns:
        Dict mit Key (tabellenname, spaltenname) oder (name, None) für einfache Bereiche.
        Value: {"blatt": str, "bereich": str, "bereich_mit_blatt": str}
    """
    index: dict[tuple[str, Optional[str]], dict[str, str]] = {}

    # 1. ws.tables (ListObjects) – alle Blätter
    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            if not hasattr(ws, "tables") or not ws.tables:
                continue
            for table_name, table in ws.tables.items():
                table_ref = getattr(table, "ref", None)
                if not table_ref:
                    continue
                table_columns = getattr(table, "tableColumns", []) or []
                # Tabellen-Gesamtbereich (ohne Spalte)
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(table_ref)
                    bereich_gesamt = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                    key_gesamt = (table_name, None)
                    if key_gesamt not in index:
                        index[key_gesamt] = {
                            "blatt": sheet_name,
                            "bereich": bereich_gesamt,
                            "bereich_mit_blatt": f"'{sheet_name}'!{bereich_gesamt}",
                        }
                except Exception:
                    pass
                # Pro Spalte
                for col in table_columns:
                    col_name = getattr(col, "name", None) or ""
                    if not col_name or str(col_name).startswith("#"):
                        continue
                    key = (table_name, str(col_name).strip())
                    if key in index:
                        continue
                    result = _spaltenbereich_aus_tabelle(
                        table_ref, table_columns, col_name, sheet_name
                    )
                    if result:
                        index[key] = result
        except Exception as e:
            logger.debug(f"Fehler beim Verarbeiten von Blatt {sheet_name}: {e}")

    # 2. wb.defined_names – globale benannte Bereiche
    try:
        def_names = getattr(wb, "defined_names", None)
        if def_names:
            # openpyxl: defined_names ist dict-like oder hat .definedName
            items = []
            if hasattr(def_names, "items"):
                items = list(def_names.items())
            elif hasattr(def_names, "definedName"):
                for defn in def_names.definedName:
                    name = getattr(defn, "name", None)
                    if name:
                        items.append((name, defn))
            for name, defn in items:
                if not name or not defn:
                    continue
                try:
                    dests = list(getattr(defn, "destinations", None) or [])
                    if not dests:
                        continue
                    sheet_title, coord = dests[0]
                    key = (str(name), None)
                    if key not in index:
                        index[key] = {
                            "blatt": sheet_title,
                            "bereich": coord,
                            "bereich_mit_blatt": f"'{sheet_title}'!{coord}",
                        }
                except Exception as e:
                    logger.debug(f"Defined name {name} nicht auflösbar: {e}")
    except Exception as e:
        logger.debug(f"defined_names fehlgeschlagen: {e}")

    # 3. ws.defined_names – blattspezifische benannte Bereiche (falls vorhanden)
    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            ws_defns = getattr(ws, "defined_names", None)
            if not ws_defns:
                continue
            items = list(ws_defns.items()) if hasattr(ws_defns, "items") else []
            for name, defn in items:
                if not name or not defn:
                    continue
                try:
                    dests = list(getattr(defn, "destinations", None) or [])
                    if not dests:
                        continue
                    sheet_title, coord = dests[0]
                    key = (name, None)
                    if key not in index:
                        index[key] = {
                            "blatt": sheet_title,
                            "bereich": coord,
                            "bereich_mit_blatt": f"'{sheet_title}'!{coord}",
                        }
                except Exception as e:
                    logger.debug(f"Defined name {name} (Blatt {sheet_name}) nicht auflösbar: {e}")
        except Exception as e:
            logger.debug(f"defined_names für Blatt {sheet_name} fehlgeschlagen: {e}")

    logger.info(f"Referenz-Index: {len(index)} Einträge geladen")
    return index
